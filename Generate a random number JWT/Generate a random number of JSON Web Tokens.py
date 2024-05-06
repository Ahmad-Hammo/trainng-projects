# headers
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException 
from openpyxl import load_workbook
import pandas as pd 
import time 
import random

# set options for driver
options = Options()
options.add_argument("--ignore-certificate-errors")  
options.add_argument('--allow-running-insecure-content') 
driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.set_page_load_timeout(3)  # Timeout after 3 seconds

# Define needed Varibels  
rand = random.randint(5, 16) 
UnCheckBox = rand
count = 1

# Define a function to save data
def save_data_to_excel(data, filename='The Final Result.xlsx'):
    df = pd.DataFrame([data])
    try:
        # Load the existing workbook
        book = load_workbook(filename)
        sheet = book.active  # Assume you're working with the first sheet

        # Determine the next empty row (after the last non-empty row)
        max_row = sheet.max_row
        while max_row > 0 and all(sheet.cell(row=max_row, column=col).value is None for col in range(1, sheet.max_column + 1)):
            max_row -= 1

        # Append data frame to the Excel sheet, starting at the first empty row after last filled
        for idx, row in enumerate(df.itertuples(index=False, name=None), start=max_row + 2):
            for col, value in enumerate(row, start=1):
                sheet.cell(row=idx, column=col, value=value)

        # Save the workbook
        book.save(filename)
    except FileNotFoundError:
        # If the file does not exist, create it and save
        df.to_excel(filename, sheet_name = "Fake Users",index=False)

# log messege for number of iterations
print("this process will retry " + str(rand) + " times")

while(rand > 0):
    #________________________________________________First web page________________________________________________________________
    # Online GUID Generator website

    # attach web page 
    driver.get('https://www.guidgenerator.com/online-guid-generator.aspx?v=12') 

    # set the num of GUID to 13522
    inp = driver.find_element(By.NAME, 'txtCount') 
    inp.click()
    inp.clear()
    inp.send_keys('1')

    # uncheck Hypens checkbox
    if(UnCheckBox == rand):
        driver.find_element(By.ID, 'chkHypens').click()

    # click generate some GUID's Button
    driver.find_element(By.NAME, 'btnGenerate').click() 
    time.sleep(1)

    # get GUID from the text box
    Guid = driver.find_element(By.NAME, 'txtResults').text

    #________________________________________________Second web page_______________________________________________________________
    # Fake Name Generator website

    # attach web page 
    try:
        driver.get('https://www.fakenamegenerator.com/gen-male-ar-sp.php')
    except TimeoutException:
        driver.refresh()


    # setup the dropDown lists as we need (Random , Arabic , Tunisia)
    Select(driver.find_element(By.NAME, 'gen')).select_by_value('random')
    Select(driver.find_element(By.NAME, 'n')).select_by_value('ar')
    Select(driver.find_element(By.NAME, 'c')).select_by_value('tn')

    # Click Generate Button
    driver.find_element(By.ID, 'genbtn').click()
    time.sleep(1)

    # extract full name and seperate it to first and last
    full_name = driver.find_element(By.CLASS_NAME, 'address').find_element(By.TAG_NAME, 'h3').text
    name_parts = full_name.split()

    # the condition to handle if the first name from tow part like (al najwa) becouse we deal with arabic name
    if len(name_parts) > 2:
        first_name = " ".join(name_parts[:-1])  
        last_name = name_parts[-1] 
    else:
        first_name = name_parts[0] if len(name_parts) > 0 else ""
        last_name = name_parts[1] if len(name_parts) > 1 else ""
    first_name = name_parts[0]
    last_name = name_parts[-1] 

    # extract all nedded information
    mother_maiden_name = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Mother')]]/dd").text
    phone = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Phone')]]/dd").text
    country_code = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Country code')]]/dd").text
    birthdate = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Birthday')]]/dd").text
    company = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Company')]]/dd").text
    favorite_color = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Favorite color')]]/dd").text
    email_element = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Email Address')]]/dd").text
    email_address = email_element.split()[0]  
    password = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Password')]]/dd").text

    #________________________________________________Third web page_______________________________________________________________
    # Generate JSON Web Token (JWT)

    # attach web page 
    driver.get('http://jwtbuilder.jamiekurtz.com/?v=13')
    time.sleep(2)

    # click clear all button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[1]').click()

    # insert first name
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click() # Add One Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr/td[1]/input').send_keys('First Name')
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr/td[2]/input').send_keys(first_name)

    # insert last name
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click() # Add One Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[2]/td[1]/input').send_keys('Last Name')
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[2]/td[2]/input').send_keys(last_name)

    # insert Mother's Maiden Name    
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click() # Add One Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[3]/td[1]/input').send_keys("Mother's Maiden Name")
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[3]/td[2]/input').send_keys(mother_maiden_name)

    # insert Phone Number 
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click() # Add One Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[4]/td[1]/input').send_keys('PHONE')
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[4]/td[2]/input').send_keys(phone)

    # insert Country Code 
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click() # Add One Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[5]/td[1]/input').send_keys('Country Code')
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[5]/td[2]/input').send_keys(country_code)

    # insert Birthdate 
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click() # Add One Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[6]/td[1]/input').send_keys('Birthdate')
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[6]/td[2]/input').send_keys(birthdate)

    # insert Company 
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click() # Add One Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[7]/td[1]/input').send_keys('Company')
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[7]/td[2]/input').send_keys(company)

    # insert Favorite Color 
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click() # Add One Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[8]/td[1]/input').send_keys('Favorite Color')
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[8]/td[2]/input').send_keys(favorite_color)

    # insert Email Address
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click() # Add One Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[9]/td[1]/input').send_keys('Email Address')
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[9]/td[2]/input').send_keys(email_address)

    # insert Password
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click() # Add One Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[10]/td[1]/input').send_keys('Password')
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[10]/td[2]/input').send_keys(password)

    # insert GUID
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click() # Add One Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[11]/td[1]/input').send_keys('GUID')
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[11]/td[2]/input').send_keys(Guid)

    # Click to generate 32-bit key Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[4]/div[3]/div[3]/div/button[2]').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[4]/div[3]/div[3]/div/ul/li[1]/a').click()
    time.sleep(2)

    #  Click to Create signed JWT Button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[4]/div[3]/div[3]/div/button[1]').click()
    time.sleep(2)

    # Get the JSON and save it
    full_json = driver.find_element(By.ID, 'created-jwt').text
    Json = full_json[-32:] 

    # log all information
    print("\n-------------------------------------------")
    print("\tperson profile number : " + str(count))
    print("Full Name : ", full_name)
    print("first Name : ", first_name)
    print("last Name : ", last_name)
    print("Mother's Maiden Name :", mother_maiden_name)
    print("Phone :", phone)
    print("Country Code :", country_code)
    print("Birthdate :", birthdate)
    print("Company :", company)
    print("Favorite Color :", favorite_color)
    print("Email Address :", email_address)
    print("Password :", password)
    print("GUID : ", Guid)
    print("JTW : ", Json)
    print("-------------------------------------------")

    # Prepare data dictionary
    person_data = {
        "First Name": first_name,
        "Last Name": last_name,
        "Mother's Maiden Name": mother_maiden_name,
        "Phone": phone,
        "Country Code": country_code,
        "Birthdate": birthdate,
        "Company": company,
        "Favorite Color": favorite_color,
        "Email Address": email_address,
        "Password": password,
        "GUID": Guid,
        "JWT": Json
    }

    # Save data to Excel
    save_data_to_excel(person_data)

    # update counters
    rand -= 1
    count += 1


