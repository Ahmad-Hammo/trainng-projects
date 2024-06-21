try:
    # headers
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import Select
    from selenium.webdriver.chrome.options import Options
    from selenium.common.exceptions import TimeoutException 
    from openpyxl import load_workbook
    import pandas as pd 
    import logging
    import datetime
    import time 
    import random
    import requests

    # start time of run (to calculate the run time)
    start_time = time.time()

    # Set up logging (%(asctime)s   filemode='w')
    logging.basicConfig(filename='Automation_Logs.log', level=logging.INFO, format='%(levelname)s - %(message)s')

    logging.info(f'\t\t\tRunning at {datetime.datetime.now()}\n')

    logging.info('start set options for driver.')
    # set options for driver
    options = Options()
    options.add_argument("--ignore-certificate-errors")  
    options.add_argument('--allow-running-insecure-content') 
    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    driver.set_page_load_timeout(5)  # Timeout after 5 seconds

    # Load the configuration from Excel
    config_f = pd.read_excel('config.xlsx', index_col="Name")

    # Access configurations
    fake_name_url = config_f.at['fake_name_generator_url', 'Value']
    guid_generator_url = config_f.at['guid_generator_url', 'Value']
    jwt_builder_url = config_f.at['jwt_builder_url', 'Value']
    gender = config_f.at['Gender', 'Value']
    Name_Set = config_f.at['Name_Set', 'Value']
    Country = config_f.at['Country', 'Value']
    queueName = config_f.at['queue_name', 'Value']

     # Orchestrator API details
    orchestrator_url = config_f.at['orchestrator_url', 'Value']
    tenant_name = config_f.at['tenant_name', 'Value']
    auth_token = "rt_E52014ECAF2A8A4DA515CA88FDD14541554FD7EE77FA6839009D6488C445F280-1"  
    folder_id = "5218441"  

    # Headers for the API request
    headers = {
        "Content-Type": "application/json",
        "X-UIPATH-TenantName": tenant_name,
        "Authorization": f"Bearer {auth_token}",
        "X-UIPATH-OrganizationUnitId": folder_id
            }
    
    # Function to fetch asset value
    def fetch_asset_value(asset_name):
        response = requests.get(f"{orchestrator_url}/odata/Assets?$filter=Name eq '{asset_name}'", headers=headers)
        if response.status_code == 200:
            asset_value = response.json()['value'][0]['Value']
            return int(asset_value)
        else:
            raise Exception(f"Failed to fetch asset '{asset_name}': {response.text}")

    from_ = fetch_asset_value('start range of random number')
    to = fetch_asset_value('end range of random number')

    # Function to add a transaction item to the queue
    def add_transaction_to_queue(guid):
        queue_name = queueName
        endpoint = f"{orchestrator_url}/odata/Queues/UiPathODataSvc.AddQueueItem"
        
        payload = {
            "itemData": {
                "Name": queue_name,
                "SpecificContent": {
                    "GUID": guid
                }
            }
        }
        
        response = requests.post(endpoint, json=payload, headers=headers)
        
        if response.status_code == 201:
            logging.info(f"Successfully added GUID {guid} to queue.")
        else:
            logging.error(f"Failed to add GUID {guid} to queue: {response.text}")

    # Function to fetch a transaction item from the queue
    def fetch_transaction_from_queue():
        endpoint = f"{orchestrator_url}/odata/QueueItems?$filter=Status eq 'New'&$top=1"
        
        response = requests.get(endpoint, headers=headers)
        
        if response.status_code == 200 and response.json()['value']:
            transaction = response.json()['value'][0]
            return transaction
        else:
            raise Exception(f"Failed to fetch transaction from queue: {response.text}")

    # Define a function to save data
    def save_data_to_excel(data, filename='The Final Result.xlsx'):
        df = pd.DataFrame([data])
        try:
            # Load the existing workbook
            book = load_workbook(filename)
            sheet = book.active  # Assume you're working with the first sheet

            # Determine the next empty row (after the last non-empty row)
            max_row = sheet.max_row
            while max_row > 0 and all(sheet.cell(row=max_row, column=col).value is None for col in range(1, sheet.max_column + 1)):
                max_row -= 1

            # Append data frame to the Excel sheet, starting at the first empty row after last filled
            for idx, row in enumerate(df.itertuples(index=False, name=None), start=max_row + 2):
                for col, value in enumerate(row, start=1):
                    sheet.cell(row=idx, column=col, value=value)

            # Save the workbook
            book.save(filename)
            logging.info(f'Successfully saved data for {data["First Name"]} {data["Last Name"]} to Excel.')

        except FileNotFoundError:
            # If the file does not exist, create it and save
            logging.error(f'File not found: {filename}. Creating a new file.')
            df.to_excel(filename, sheet_name = "Fake Users",index=False)
            logging.info(f'Successfully saved data for {data["First Name"]} {data["Last Name"]} to Excel.')

        except Exception as e:
            logging.error(f'Failed to save data to Excel: {str(e)}')

    # Define needed Varibels  
    rand = random.randint(from_, to) 
    UnCheckBox = rand
    count = 1

    logging.info('Start opening initial URLs.')
    # initial all applications
    try:
        driver.get(fake_name_url)
    except TimeoutException:
        logging.error('Page load timed out, trying to refresh.')
        driver.refresh()
        time.sleep(3)
        
    try:
        driver.execute_script(f"window.open('{guid_generator_url}');")
    except TimeoutException:
        logging.error('Page load timed out, trying to refresh.')
        driver.refresh()
        time.sleep(3)

    try:
        driver.execute_script(f"window.open('{jwt_builder_url}');")
    except TimeoutException:
        logging.error('Page load timed out, trying to refresh.')
        driver.refresh()
        time.sleep(3)

    logging.info('Setting up initial configurations on each website.')
    #_____________________________________________________________________seting all applications__________________________________________________________________
    # Online GUID Generator website
    # attach web page
    driver.switch_to.window(driver.window_handles[2])
    time.sleep(2)
    # set the num of GUID to Needed
    inp = driver.find_element(By.NAME, 'txtCount') 
    inp.click()
    inp.clear()
    inp.send_keys(f'{rand}')

        # uncheck Hyphens checkbox
    if(UnCheckBox == rand):
            driver.find_element(By.ID, 'chkHyphens').click()

        # click generate some GUID's Button
    driver.find_element(By.NAME, 'btnGenerate').click() 
    time.sleep(1)

        # get GUID's from the text box
    guids = driver.find_element(By.NAME, 'txtResults').text.split()
    driver.close()

    # Add generated GUIDs to the queue
    for guid in guids:
        add_transaction_to_queue(guid)
    
    # Fake Name Generator website
    # attach web page 
    driver.switch_to.window(driver.window_handles[0])
    time.sleep(2)
    Select(driver.find_element(By.NAME, 'gen')).select_by_value(gender)
    Select(driver.find_element(By.NAME, 'n')).select_by_value(Name_Set)
    Select(driver.find_element(By.NAME, 'c')).select_by_value(Country)

    # Generate JSON Web Token (JWT)
    # attach web page 
    driver.switch_to.window(driver.window_handles[1])
    time.sleep(2)

    # click clear all button
    driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[1]').click()

    # setup needed feild
    inputs = ['First Name', 'Last Name', "Mother's Maiden Name", 'PHONE', 'Country Code', 'Birthdate', 'Company', 'Favorite Color', 'Email Address', 'Password']
    counter = 10
    selector = 1
    while ( selector <= counter ):
        driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[2]/div[3]/div[2]/p[4]/button[2]').click()
        driver.find_element(By.XPATH, f'/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[{selector}]/td[1]/input').send_keys(inputs[selector-1])
        selector += 1
    logging.info('End Setting.')

    #_________________________________________________________________________End setting_______________________________________________________________________________

    # log messege for number of iterations
    logging.info(f"The Random number Generated between {from_} and {to}: is {rand}")

    logging.info(f"this process will retry {rand} times")

    logging.info('Starting the main processing loop.')
    while(rand > 0):
        # Fetch transaction from the queue
        transaction = fetch_transaction_from_queue()

        guid = transaction['SpecificContent']['GUID']
        #transaction_id = transaction['Id']  # Use 'Id' instead of 'QueueItemId'

        #________________________________________________First_______________________________________________________________
        # Fake Name Generator website

        # attach web page 
        driver.switch_to.window(driver.window_handles[0])
        
        # Click Generate Button
        driver.find_element(By.ID, 'genbtn').click()
        time.sleep(1)

        # extract full name and separate it to first and last
        full_name = driver.find_element(By.CLASS_NAME, 'address').find_element(By.TAG_NAME, 'h3').text
        name_parts = full_name.split()

        # the condition to handle if the first name from two part like (al najwa) because we deal with arabic name
        if len(name_parts) > 2:
            first_name = " ".join(name_parts[:-1])  
            last_name = name_parts[-1] 
        else:
            first_name = name_parts[0] if len(name_parts) > 0 else ""
            last_name = name_parts[1] if len(name_parts) > 1 else ""
        first_name = name_parts[0]
        last_name = name_parts[-1] 

    
        # extract all needed information
        mother_maiden_name = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Mother')]]/dd").text
        phone = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Phone')]]/dd").text
        country_code = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Country code')]]/dd").text
        birthdate = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Birthday')]]/dd").text
        company = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Company')]]/dd").text
        favorite_color = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Favorite color')]]/dd").text
        email_element = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Email Address')]]/dd").text
        email_address = email_element.split()[0]  
        password = driver.find_element(By.XPATH, "//dl[dt[contains(text(), 'Password')]]/dd").text


        #________________________________________________Second_______________________________________________________________
        # Generate JSON Web Token (JWT)

        # attach web page 
        driver.switch_to.window(driver.window_handles[1])

        inputs2 = [first_name, last_name, mother_maiden_name, phone, country_code, birthdate, company, favorite_color, email_address, password]

        # insert informations
        coun = 1
        while(coun < 11 ):
            driver.find_element(By.XPATH, f'/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[{coun}]/td[2]/input').clear()
            driver.find_element(By.XPATH, f'/html/body/div[2]/div/form/div[2]/div[3]/div[1]/table/tbody/tr[{coun}]/td[2]/input').send_keys(inputs2[coun-1])
            coun += 1 

        # Click to key Text box 
        driver.find_element(By.XPATH, '//*[@id="key"]').click()
        driver.find_element(By.XPATH, '//*[@id="key"]').clear()
        driver.find_element(By.XPATH, '//*[@id="key"]').send_keys(guid)
        time.sleep(1)

        #  Click to Create signed JWT Button
        driver.find_element(By.XPATH, '/html/body/div[2]/div/form/div[4]/div[3]/div[3]/div/button[1]').click()
        time.sleep(2)

        # Get the JSON and save it
        Json = driver.find_element(By.ID, 'created-jwt').text
        

        # log all information
        logging.info("\n-------------------------------------------")
        logging.info(f"iteration number : {count}")
        logging.info(f"Full Name : {full_name}")
        logging.info(f"first Name : {first_name}")
        logging.info(f"last Name : {last_name}")
        logging.info(f"Mother's Maiden Name : {mother_maiden_name}")
        logging.info(f"Phone : {phone}")
        logging.info(f"Country Code : {country_code}")
        logging.info(f"Birthdate : {birthdate}")
        logging.info(f"Company : {company}")
        logging.info(f"Favorite Color : {favorite_color}")
        logging.info(f"Email Address : {email_address}:")
        logging.info(f"Password : {password}")
        logging.info(f"GUID : {guid}")
        logging.info(f"JTW :{Json} ")
        logging.info("\n-------------------------------------------")

        # Prepare data dictionary
        person_data = {
            "First Name": first_name,
            "Last Name": last_name,
            "Mother's Maiden Name": mother_maiden_name,
            "Phone": phone,
            "Country Code": country_code,
            "Birthdate": birthdate,
            "Company": company,
            "Favorite Color": favorite_color,
            "Email Address": email_address,
            "Password": password,
            "GUID": guid,
            "JWT": Json
        }

        # Save data to Excel
        save_data_to_excel(person_data)

        # update counters
        rand -= 1
        count += 1
        logging.info(f'Completed iteration {count-1}.')
    
    run_time = time.time() - start_time
    logging.info('Ending the main processing loop.')
    logging.info(f"Run completed. Total time taken: {run_time:.2f} seconds.\n.\n.\n.\n.\n.")
    
except Exception as e:
    run_time = time.time() - start_time
    logging.warning(f"Run stop (Error happened). Total time taken: {run_time:.2f} seconds.")
    logging.error(f'Unexpected error occurred: {str(e)}\n.\n.\n.\n.\n.')


